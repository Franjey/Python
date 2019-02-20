import PyPDF2
import os
import re
from openpyxl import Workbook, load_workbook

'''
Lógica:
1. Fijarse si la factura es en dolares o pesos
2. Fijarse el tipo/codigo de la factura: Factura A o B, y codigo 01 o 06
'''
i = 1
fila_excel = 1

importe_list = []
facturas_usd = []
facturas_pesos = []

# ----- Excel -----
wb = load_workbook("Contenedor.xlsx")
ws = wb.active
sheet = wb["Sheet1"]


def invoice_path(invoice_name):
    script_path = os.path.abspath(__file__)
    path_list = script_path.split(os.sep)
    script_directory = path_list[0:len(path_list)-1]
    rel_path = "Archivos/" + invoice_name
    path = "/".join(script_directory) + "/" + rel_path
    return path

# ----- Distinción entre facturas en Dólares y Pesos, y armado de listas -----
for root, dirs, files in os.walk('C:/Users/FRANCISCODIEGOJurado/PycharmProjects/IBM_SQL/Archivos/'):
    for file in files:
        if file.endswith('.pdf'):
            path = os.path.join(root, file)
            pdf_file = PyPDF2.PdfFileReader(open(path, 'rb'))

            #   Si el PDF está encriptado
            if pdf_file.isEncrypted:
                pdf_file.decrypt('')

            pdf_text = pdf_file.getPage(1).extractText()

            for c in range(len(pdf_text)):
                if pdf_text[c:c+3] == "USD":
                    facturas_usd.append(file)
                else:
                    facturas_pesos.append(file)

# ----- Remoción de duplicados ------
# Remover facturas en dólares, de la lista de facturas en pesos
for i in facturas_pesos:
    for x in facturas_usd:
        if x in facturas_pesos:
            facturas_pesos.remove(x)
# print(list(set(facturas_pesos)))

# ----- Inicio: Facturas en Dólares -----
for i in list(set(facturas_usd)):
    usd_file = PyPDF2.PdfFileReader(invoice_path(i),'rb')

    if usd_file.isEncrypted:
        usd_file.decrypt('')

    usd_file_text = usd_file.getPage(1).extractText()

    patron_3 = re.compile('consumidor\"')
    matches_4 = patron_3.finditer(usd_file_text)

    for match in matches_4:
        start = match.end()

    patron_4 = re.compile('\:  \$El')
    matches_5 = patron_4.finditer(usd_file_text)

    for match in matches_5:
        end_usd = match.start()

    patron_5 = re.compile('consignado de ')
    matches_6 = patron_5.finditer(usd_file_text)

    for match in matches_6:
        tipo_de_cambio = float(usd_file_text[match.end():match.end() + 5])

    importe = float(usd_file_text[start:end_usd].replace(',', '.'))
    importe_usd = importe / tipo_de_cambio
    importe_list.append(importe_usd)
    ws.cell(fila_excel, 1, importe_usd)

    xf = re.split("/", invoice_path(i))
    xf = xf[-1]
    ws.cell(fila_excel, 2, xf)
    fila_excel = fila_excel + 1

# ----- Inicio: Facturas en Pesos -----
for i in list(set(facturas_pesos)):
    pesos_file = PyPDF2.PdfFileReader(invoice_path(i), 'rb')

    if pesos_file.isEncrypted:
        pesos_file.decrypt('')

    pesos_file_text = pesos_file.getPage(1).extractText()

    for char_index in range(len(pesos_file_text)):
        if pesos_file_text[char_index:char_index + 7] == "COD. 01":
            patron = re.compile(r'Descripción')
            matches_2 = patron.finditer(pesos_file_text)

            for match in matches_2:
                importe_end = match.start()

            patron_2 = re.compile(r'\d%\d')
            matches_3 = patron_2.finditer(pesos_file_text)
            for match in matches_3:
                importe_start = match.end()

            importe_pesos = pesos_file_text[importe_start - 1:importe_end]
            importe_list.append(importe_pesos)
            ws.cell(fila_excel, 1, importe_pesos)

            xf = re.split("/", invoice_path(i))
            xf = xf[-1]
            ws.cell(fila_excel, 2, xf)

            fila_excel = fila_excel + 1

        elif pesos_file_text[char_index:char_index + 7] == "COD. 06":
                x = 0
                patron = re.compile(r'Subtotal:')
                matches_2 = patron.finditer(pesos_file_text)

                for match in matches_2:
                    importe_end = match.start()

                for i in reversed(range(len(pesos_file_text[:importe_end]))):
                    if pesos_file_text[i] == ",":
                        x += 1
                        if x == 2:
                            importe = pesos_file_text[i + 3:importe_end]
                            importe_list.append(importe)
                            ws.cell(fila_excel, 1, importe)
                            xf = re.split("/", invoice_path(i))
                            xf = xf[-1]
                            ws.cell(fila_excel, 2, xf)
                            fila_excel = fila_excel + 1

# ----- Si hay uno de los montos que no es float, debemos convertir en Strings,
# reemplazar las comas con puntos, y transformarlos en Floats
importe_list = [str(i).replace(',','.') for i in importe_list]
importe_list = [float(i) for i in importe_list]
print(importe_list)

wb.save("Contenedor.xlsx")


